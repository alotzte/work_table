import openpyxl

# Открываем файл Excel
workbook = openpyxl.load_workbook('output.xlsx')

# Выбираем нужный лист в файле
sheet = workbook['Sheet1']

# Проверяем каждую ячейку в колонке A на пустоту
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=2):
    cell = row[0]
    if cell.value is None:
        print(f'Ячейка {cell.coordinate} пустая')

# Закрываем файл Excel
workbook.close()
